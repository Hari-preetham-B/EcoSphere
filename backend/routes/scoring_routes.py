import io
import csv
from datetime import datetime, date as date_type
from flask import Blueprint, jsonify, request, make_response, send_file, g
from sqlalchemy import func
from database import db
from models import (
    Department, UserProfile, CarbonTransaction, SustainabilityGoal,
    CSRActivity, CSRParticipation, TrainingCompletion,
    ESGPolicy, PolicyAcknowledgement, Audit, ComplianceIssue, Category
)
from auth import token_required, require_role

scoring_bp = Blueprint('scoring', __name__)

# Try imports for PDF and Excel export
try:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ─── Scoring Engine Helper ────────────────────────────────────────────────────

def compute_department_scores(dept_id, weights=None):
    """
    Computes Environmental, Social, Governance, and Total Scores for a given department.
    All scores are scaled 0 - 100.
    """
    if weights is None:
        weights = {'env': 0.40, 'soc': 0.30, 'gov': 0.30}

    dept = Department.query.get(dept_id)
    if not dept:
        return None

    dept_users = UserProfile.query.filter_by(department=dept.name).all()
    dept_user_ids = [u.id for u in dept_users]
    dept_user_count = max(1, len(dept_users))

    # 1. Environmental Score Components
    # (a) Relative Emissions Benchmark
    dept_tx_count = CarbonTransaction.query.filter_by(department_id=dept_id).count()
    dept_tx_sum = db.session.query(func.coalesce(func.sum(CarbonTransaction.co2e), 0.0))\
        .filter(CarbonTransaction.department_id == dept_id).scalar() or 0.0
    dept_avg_emissions = float(dept_tx_sum / max(1, dept_tx_count)) if dept_tx_count > 0 else 0.0

    all_tx_count = CarbonTransaction.query.count()
    all_tx_sum = db.session.query(func.coalesce(func.sum(CarbonTransaction.co2e), 0.0)).scalar() or 0.0
    company_avg_emissions = float(all_tx_sum / max(1, all_tx_count)) if all_tx_count > 0 else 0.0

    if all_tx_count == 0:
        env_benchmark_score = 100.0
    elif dept_tx_count == 0:
        env_benchmark_score = 75.0  # Default neutral score for no transactions
    elif dept_avg_emissions <= company_avg_emissions:
        ratio = (dept_avg_emissions / max(0.001, company_avg_emissions))
        env_benchmark_score = round(min(100.0, 50.0 + 50.0 * (1.0 - ratio)), 1)
    else:
        ratio = (company_avg_emissions / max(0.001, dept_avg_emissions))
        env_benchmark_score = round(max(0.0, 50.0 * ratio), 1)

    # (b) Sustainability Goals Progress
    dept_goals = SustainabilityGoal.query.filter(
        (SustainabilityGoal.department_id == dept_id) | (SustainabilityGoal.department_id.is_(None))
    ).all()

    if not dept_goals:
        env_goal_score = 100.0
    else:
        goal_scores = []
        for g in dept_goals:
            tx_q = db.session.query(func.coalesce(func.sum(CarbonTransaction.co2e), 0.0))\
                .filter(CarbonTransaction.date <= g.deadline)
            if g.department_id:
                tx_q = tx_q.filter(CarbonTransaction.department_id == g.department_id)
            actual_co2e = float(tx_q.scalar() or 0.0)
            target = float(g.target_value) if g.target_value else 100.0

            # Direction: emission cap ceiling (lower actual emissions is better)
            # 0 actual = 100%, actual == target = 50%, actual >= 2*target = 0%
            g_score = max(0.0, min(100.0, 100.0 * (1.0 - (actual_co2e / (2.0 * target)))))
            goal_scores.append(g_score)
        env_goal_score = round(sum(goal_scores) / max(1, len(goal_scores)), 1)

    env_score = round(0.5 * env_benchmark_score + 0.5 * env_goal_score, 1)

    # 2. Social Score Components
    # (a) CSR Participation Rate (department employee normalized)
    csr_part_users = db.session.query(CSRParticipation.user_id)\
        .filter(CSRParticipation.status == 'Approved')\
        .filter(CSRParticipation.user_id.in_(dept_user_ids)).distinct().count() if dept_user_ids else 0
    csr_score = round(min(100.0, (csr_part_users / dept_user_count) * 100.0), 1)

    # (b) Training Completion Rate
    dept_completions = TrainingCompletion.query.filter(
        TrainingCompletion.user_id.in_(dept_user_ids)
    ).all() if dept_user_ids else []

    if not dept_completions:
        training_score = 100.0
    else:
        total_assigned = len(dept_completions)
        total_completed = sum(1 for c in dept_completions if c.status == 'Completed')
        training_score = round((total_completed / max(1, total_assigned)) * 100.0, 1)

    soc_score = round(0.5 * csr_score + 0.5 * training_score, 1)

    # 3. Governance Score Components
    # (a) Policy Acknowledgement Rate (strictly filtered to applicable policies)
    applicable_policies = ESGPolicy.query.filter(
        (ESGPolicy.department_id == dept_id) | (ESGPolicy.department_id.is_(None)),
        ESGPolicy.status == 'Active'
    ).all()
    applicable_policy_ids = [p.id for p in applicable_policies]
    expected_acks = max(1, len(applicable_policies) * dept_user_count)

    actual_acks = 0
    if applicable_policy_ids and dept_user_ids:
        actual_acks = PolicyAcknowledgement.query.filter(
            PolicyAcknowledgement.policy_id.in_(applicable_policy_ids),
            PolicyAcknowledgement.user_id.in_(dept_user_ids),
            PolicyAcknowledgement.status == 'Acknowledged'
        ).count()

    gov_ack_rate = round(min(100.0, (actual_acks / expected_acks) * 100.0), 1)

    # (b) Compliance Issues & Scaled Penalty
    dept_audits = Audit.query.filter_by(department_id=dept_id).all()
    dept_audit_ids = [a.id for a in dept_audits]

    if not dept_audit_ids:
        gov_issue_score = 100.0
        overdue_count = 0
    else:
        dept_issues = ComplianceIssue.query.filter(ComplianceIssue.audit_id.in_(dept_audit_ids)).all()
        if not dept_issues:
            gov_issue_score = 100.0
            overdue_count = 0
        else:
            resolved_count = sum(1 for i in dept_issues if i.status == 'Resolved')
            base_res_rate = (resolved_count / max(1, len(dept_issues))) * 100.0
            overdue_count = sum(1 for i in dept_issues if i.is_overdue)
            penalty_factor = 1.0 / (1.0 + 0.25 * overdue_count)
            gov_issue_score = round(max(0.0, min(100.0, base_res_rate * penalty_factor)), 1)

    gov_score = round(0.5 * gov_ack_rate + 0.5 * gov_issue_score, 1)

    # 4. Total Weighted Department Score
    total_score = round(
        weights['env'] * env_score +
        weights['soc'] * soc_score +
        weights['gov'] * gov_score,
        1
    )

    return {
        'department_id': dept.id,
        'department_name': dept.name,
        'department_code': dept.code,
        'head': dept.head,
        'employee_count': dept_user_count,
        'scores': {
            'environmental': env_score,
            'social': soc_score,
            'governance': gov_score,
            'total': total_score,
        },
        'formula_breakdown': {
            'env_benchmark_score': env_benchmark_score,
            'env_goal_score': env_goal_score,
            'dept_avg_emissions': round(dept_avg_emissions, 2),
            'company_avg_emissions': round(company_avg_emissions, 2),
            'csr_score': csr_score,
            'training_score': training_score,
            'gov_ack_rate': gov_ack_rate,
            'gov_issue_score': gov_issue_score,
            'overdue_issues_count': overdue_count,
        }
    }


def seed_multi_department_scoring_data():
    """
    Seeds varied data across 3 departments so scores are visually distinct and testable.
    """
    try:
        from models import (
            Department, UserProfile, CarbonTransaction, Category,
            EmissionFactor, CSRActivity, CSRParticipation, TrainingCompletion,
            ESGPolicy, PolicyAcknowledgement, Audit, ComplianceIssue
        )

        # 1. Create 3 distinct departments if they don't exist
        depts_def = [
            {'name': 'Sustainability & Operations', 'code': 'SO-01', 'head': 'Jane Doe'},
            {'name': 'Engineering & Technology', 'code': 'ET-02', 'head': 'Alex Chen'},
            {'name': 'Human Resources & Facilities', 'code': 'HR-03', 'head': 'Maria Garcia'},
        ]
        dept_objs = {}
        for d in depts_def:
            obj = Department.query.filter_by(name=d['name']).first()
            if not obj:
                obj = Department(name=d['name'], code=d['code'], head=d['head'], status='Active')
                db.session.add(obj)
                db.session.flush()
            dept_objs[d['name']] = obj

        so_dept = dept_objs['Sustainability & Operations']
        et_dept = dept_objs['Engineering & Technology']
        hr_dept = dept_objs['Human Resources & Facilities']
        db.session.commit()

        # Update or create users assigned to departments
        user1 = UserProfile.query.filter_by(role='Admin').first()
        if user1:
            user1.department = so_dept.name

        emp_users = UserProfile.query.filter(UserProfile.role == 'Employee').all()
        if len(emp_users) >= 1:
            emp_users[0].department = et_dept.name
        if len(emp_users) >= 2:
            emp_users[1].department = hr_dept.name

        db.session.commit()

        # 2. Seed Carbon Transactions (varied emissions per dept)
        if CarbonTransaction.query.count() < 4:
            cat = Category.query.first() or Category(name='Energy', type='Scope 1', status='Active')
            ef = EmissionFactor.query.first() or EmissionFactor(name='Electricity Grid', category_id=1, factor=0.5, unit='kgCO2e/kWh', source='DEFRA')
            if not cat.id: db.session.add(cat); db.session.flush()
            if not ef.id: db.session.add(ef); db.session.flush()

            tx1 = CarbonTransaction(department_id=so_dept.id, source='Electricity', quantity=100, emission_factor_id=ef.id, co2e=50.0, date=date_type(2026, 6, 1))
            tx2 = CarbonTransaction(department_id=et_dept.id, source='Natural Gas', quantity=800, emission_factor_id=ef.id, co2e=400.0, date=date_type(2026, 6, 10))
            tx3 = CarbonTransaction(department_id=hr_dept.id, source='Fleet', quantity=200, emission_factor_id=ef.id, co2e=100.0, date=date_type(2026, 6, 15))
            db.session.add_all([tx1, tx2, tx3])
            db.session.commit()

        # 3. Seed Training Completions
        if TrainingCompletion.query.count() == 0:
            admin_user = UserProfile.query.first()
            if admin_user:
                tc1 = TrainingCompletion(user_id=admin_user.id, training_name='ISO 14001 Environmental Safety', status='Completed', completion_date=date_type(2026, 5, 10))
                tc2 = TrainingCompletion(user_id=admin_user.id, training_name='Green Software Engineering', status='Enrolled', completion_date=date_type(2026, 6, 5))
                db.session.add_all([tc1, tc2])
                db.session.commit()

        # 4. Seed Compliance Audits & Issues (varied overdue status)
        admin_user = UserProfile.query.first()
        if admin_user:
            aud_et = Audit.query.filter_by(department_id=et_dept.id).first()
            if not aud_et:
                aud_et = Audit(title='Engineering Safety Audit', department_id=et_dept.id, auditor_name='SafetyCorp', audit_date=date_type(2026, 4, 1), status='Completed', created_by=admin_user.id)
                db.session.add(aud_et)
                db.session.commit()

                iss_overdue = ComplianceIssue(
                    audit_id=aud_et.id, severity='High', description='Server room emergency power battery expired',
                    owner_id=admin_user.id, due_date=date_type(2026, 5, 1), status='Open'
                )
                iss_resolved = ComplianceIssue(
                    audit_id=aud_et.id, severity='Low', description='Update server room safety signage',
                    owner_id=admin_user.id, due_date=date_type(2026, 6, 1), status='Resolved'
                )
                db.session.add_all([iss_overdue, iss_resolved])
                db.session.commit()

        print("Multi-department scoring seed data initialized.")
    except Exception as e:
        db.session.rollback()
        print(f"Scoring seed note: {e}")


# ─── API Routes ───────────────────────────────────────────────────────────────

@scoring_bp.route('/departments', methods=['GET'])
@token_required
def get_department_scores():
    from models import Setting
    w_env = float((Setting.query.filter_by(key='weight_env').first() or type('obj', (), {'value': '0.40'})).value)
    w_soc = float((Setting.query.filter_by(key='weight_soc').first() or type('obj', (), {'value': '0.30'})).value)
    w_gov = float((Setting.query.filter_by(key='weight_gov').first() or type('obj', (), {'value': '0.30'})).value)
    weights = {'env': w_env, 'soc': w_soc, 'gov': w_gov}

    depts = Department.query.all()
    results = []
    for dept in depts:
        score_data = compute_department_scores(dept.id, weights)
        if score_data:
            results.append(score_data)

    results.sort(key=lambda x: x['scores']['total'], reverse=True)
    for rank, item in enumerate(results, 1):
        item['rank'] = rank

    return jsonify(results), 200


@scoring_bp.route('/organization', methods=['GET'])
@token_required
def get_organization_score():
    from models import Setting
    w_env = float((Setting.query.filter_by(key='weight_env').first() or type('obj', (), {'value': '0.40'})).value)
    w_soc = float((Setting.query.filter_by(key='weight_soc').first() or type('obj', (), {'value': '0.30'})).value)
    w_gov = float((Setting.query.filter_by(key='weight_gov').first() or type('obj', (), {'value': '0.30'})).value)
    weights = {'env': w_env, 'soc': w_soc, 'gov': w_gov}

    depts = Department.query.all()
    dept_scores = []
    env_scores = []
    soc_scores = []
    gov_scores = []

    for dept in depts:
        sd = compute_department_scores(dept.id, weights)
        if sd:
            dept_scores.append(sd['scores']['total'])
            env_scores.append(sd['scores']['environmental'])
            soc_scores.append(sd['scores']['social'])
            gov_scores.append(sd['scores']['governance'])

    overall_esg = round(sum(dept_scores) / max(1, len(dept_scores)), 1) if dept_scores else 100.0
    overall_env = round(sum(env_scores) / max(1, len(env_scores)), 1) if env_scores else 100.0
    overall_soc = round(sum(soc_scores) / max(1, len(soc_scores)), 1) if soc_scores else 100.0
    overall_gov = round(sum(gov_scores) / max(1, len(gov_scores)), 1) if gov_scores else 100.0

    # Trend data (simulated monthly trend based on actual current scores)
    trend = [
        {'month': 'Jan', 'score': max(0, round(overall_esg - 5.2, 1))},
        {'month': 'Feb', 'score': max(0, round(overall_esg - 3.8, 1))},
        {'month': 'Mar', 'score': max(0, round(overall_esg - 2.1, 1))},
        {'month': 'Apr', 'score': max(0, round(overall_esg - 1.0, 1))},
        {'month': 'May', 'score': max(0, round(overall_esg - 0.4, 1))},
        {'month': 'Jun', 'score': overall_esg},
    ]

    return jsonify({
        'overall_esg_score': overall_esg,
        'pillar_averages': {
            'environmental': overall_env,
            'social': overall_soc,
            'governance': overall_gov,
        },
        'weights': weights,
        'department_count': len(dept_scores),
        'trend': trend,
    }), 200


# ─── Fixed Reports API ────────────────────────────────────────────────────────

@scoring_bp.route('/reports/fixed', methods=['GET'])
@token_required
def get_fixed_report():
    report_type = request.args.get('type', 'summary')
    module_filter = request.args.get('module')  # Environmental, Social, Governance, Gamification
    dept_id = request.args.get('department_id')
    user_id = request.args.get('user_id')
    challenge_id = request.args.get('challenge_id')
    category_id = request.args.get('category_id')
    date_from_str = request.args.get('date_from')
    date_to_str = request.args.get('date_to')

    date_from = date_type.fromisoformat(date_from_str) if date_from_str else None
    date_to = date_type.fromisoformat(date_to_str) if date_to_str else None

    # Determine effective report target based on module_filter or report_type
    target_module = module_filter.lower() if module_filter else report_type.lower()

    if target_module in ['environmental', 'env']:
        q = CarbonTransaction.query
        if dept_id: q = q.filter_by(department_id=int(dept_id))
        if category_id: q = q.filter_by(category_id=int(category_id))
        if date_from: q = q.filter(CarbonTransaction.date >= date_from)
        if date_to: q = q.filter(CarbonTransaction.date <= date_to)
        txs = q.order_by(CarbonTransaction.date.desc()).all()
        return jsonify({
            'report_name': 'Environmental Emissions & Carbon Footprint Report',
            'generated_at': datetime.utcnow().isoformat(),
            'record_count': len(txs),
            'data': [t.to_dict() for t in txs]
        }), 200

    elif target_module in ['social', 'gamification']:
        q = CSRParticipation.query
        if user_id: q = q.filter_by(user_id=user_id)
        if challenge_id:
            q = q.filter_by(activity_id=int(challenge_id))
        parts = q.order_by(CSRParticipation.registered_at.desc()).all()
        return jsonify({
            'report_name': 'Social & Gamification Engagement Report',
            'generated_at': datetime.utcnow().isoformat(),
            'record_count': len(parts),
            'data': [p.to_dict() for p in parts]
        }), 200

    elif target_module in ['governance', 'gov']:
        q = ComplianceIssue.query
        if user_id: q = q.filter_by(owner_id=user_id)
        if date_from: q = q.filter(ComplianceIssue.due_date >= date_from)
        if date_to: q = q.filter(ComplianceIssue.due_date <= date_to)
        issues = q.order_by(ComplianceIssue.due_date.asc()).all()
        return jsonify({
            'report_name': 'Governance Audits & Compliance Issues Report',
            'generated_at': datetime.utcnow().isoformat(),
            'record_count': len(issues),
            'data': [i.to_dict() for i in issues]
        }), 200

    else:  # summary / overall
        depts = Department.query.all()
        if dept_id:
            depts = [d for d in depts if d.id == int(dept_id)]
        scores = [compute_department_scores(d.id) for d in depts if compute_department_scores(d.id)]
        return jsonify({
            'report_name': 'Executive ESG Performance Summary Report',
            'generated_at': datetime.utcnow().isoformat(),
            'record_count': len(scores),
            'data': scores
        }), 200


# ─── Report Export (PDF, Excel, CSV) ──────────────────────────────────────────

@scoring_bp.route('/reports/export', methods=['POST'])
@token_required
def export_report():
    data = request.get_json() or {}
    export_format = data.get('format', 'csv').lower()
    report_type = data.get('report_type', 'summary')
    rows = data.get('rows', [])

    filename_base = f"EcoSphere_{report_type}_Report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}"

    if export_format == 'csv':
        output = io.StringIO()
        if rows and isinstance(rows[0], dict):
            writer = csv.DictWriter(output, fieldnames=list(rows[0].keys()))
            writer.writeheader()
            for r in rows:
                writer.writerow(r)
        else:
            output.write("No data available\n")

        res = make_response(output.getvalue())
        res.headers["Content-Disposition"] = f"attachment; filename={filename_base}.csv"
        res.headers["Content-Type"] = "text/csv; charset=utf-8"
        return res

    elif export_format == 'excel':
        if not HAS_OPENPYXL:
            return jsonify({'error': 'openpyxl is not installed on the server.'}), 500

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ESG Report"

        # Header styling
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')

        if rows and isinstance(rows[0], dict):
            headers = list(rows[0].keys())
            ws.append([h.replace('_', ' ').title() for h in headers])
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill

            for r in rows:
                ws.append([str(v) if v is not None else '' for v in r.values()])
        else:
            ws.append(["No data available"])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"{filename_base}.xlsx"
        )

    elif export_format == 'pdf':
        if not HAS_REPORTLAB:
            return jsonify({'error': 'reportlab is not installed on the server.'}), 500

        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter, leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36)
        styles = getSampleStyleSheet()
        elements = []

        title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#0F172A'), spaceAfter=8)
        elements.append(Paragraph(f"EcoSphere ESG Platform — {report_type.upper()} Report", title_style))
        elements.append(Paragraph(f"Generated on: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", styles['Normal']))
        elements.append(Spacer(1, 12))

        if rows and isinstance(rows[0], dict):
            keys = list(rows[0].keys())[:6]  # Limit to 6 columns for PDF fit
            table_data = [[k.replace('_', ' ').title() for k in keys]]
            for r in rows:
                table_data.append([str(r.get(k, ''))[:30] for k in keys])

            t = Table(table_data)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
            ]))
            elements.append(t)
        else:
            elements.append(Paragraph("No data available", styles['Normal']))

        doc.build(elements)
        output.seek(0)
        return send_file(
            output,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=f"{filename_base}.pdf"
        )

    return jsonify({'error': 'Unsupported format'}), 400
